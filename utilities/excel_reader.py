from utilities.logger import *
from openpyxl import load_workbook
import os


class ExcelReader:
    """
    A reusable class to read Excel data — any sheet, row, column, or cell.
    """
    def __init__(self, file_path):
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"\nExcel file not found: {file_path}")

        self.file_path = file_path
        self.workbook = load_workbook(file_path)
        logger.info(f"\nWorkbook loaded successfully: {os.path.basename(file_path)}")

    def get_sheet(self, sheet_name=None):
        """Get a sheet object. If not provided, use the active sheet."""
        if sheet_name:
            if sheet_name not in self.workbook.sheetnames:
                raise ValueError(f"\nSheet '{sheet_name}' not found in workbook.")
            return self.workbook[sheet_name]
        return self.workbook.active

    def read_entire_sheet(self, sheet_name=None):
        """Return all rows from a sheet as a list of lists."""
        sheet = self.get_sheet(sheet_name)
        data = [[cell.value for cell in row] for row in sheet.iter_rows()]
        return data

    def read_row(self, row_num, sheet_name=None):
        """Read a specific row (1-based index)."""
        sheet = self.get_sheet(sheet_name)
        return [cell.value for cell in sheet[row_num]]

    def read_column(self, col_num, sheet_name=None):
        """Read a specific column (1-based index)."""
        sheet = self.get_sheet(sheet_name)
        return [cell.value for cell in sheet.iter_cols(min_col=col_num, max_col=col_num, values_only=True)][0]

    def read_cell(self, row_num, col_num, sheet_name=None):
        """Read a specific cell value."""
        sheet = self.get_sheet(sheet_name)
        return sheet.cell(row=row_num, column=col_num).value

    def close(self):
        """Close the workbook."""
        self.workbook.close()
        print("\nWorkbook closed.")



# ----------------------------------------
# Example Usage
# ----------------------------------------
# if __name__ == "__main__":
#     # Replace with your actual file path
#     file_path = "TestData/Login_Page.xlsx"
#
#     excel = ExcelReader(file_path)
#
#     print("\nEntire Sheet Data:")
#     for row in excel.read_entire_sheet("Sheet1"):
#         print(row)
#
#     print("\nRow 2 Data:")
#     print(excel.read_row(2, "Sheet1"))
#
#     print("\nColumn 1 Data:")
#     print(excel.read_column(1, "Sheet1"))
#
#     print("\nCell (2, 3) Data:")
#     print(excel.read_cell(2, 3, "Sheet1"))
#
#     excel.close()




